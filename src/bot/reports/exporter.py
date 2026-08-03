"""
Report Exporters.

Generates downloadable files (CSV, Excel) containing group statistics
and leaderboards for administrators.
"""

import csv
import io
from collections.abc import Sequence

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from bot.models.member import Member


class Exporter:
    """Handles generating downloadable data exports."""

    @staticmethod
    def generate_leaderboard_csv(members: Sequence[Member]) -> bytes:
        """
        Generate a CSV byte stream containing the leaderboard.
        """
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow([
            "Rank", 
            "User ID", 
            "First Name", 
            "Username", 
            "Total Invited", 
            "Active Invited", 
            "Join Date"
        ])
        
        # Rows
        for rank, member in enumerate(members, start=1):
            writer.writerow([
                rank,
                member.user_id,
                member.user.first_name,
                f"@{member.user.username}" if member.user.username else "",
                member.total_invited,
                member.active_invited,
                member.joined_at.strftime("%Y-%m-%d %H:%M:%S") if member.joined_at else ""
            ])
            
        return output.getvalue().encode("utf-8")

    @staticmethod
    def generate_leaderboard_excel(members: Sequence[Member]) -> bytes:
        """
        Generate a styled Excel (.xlsx) byte stream containing the leaderboard.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Invite Leaderboard"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        
        # Write headers
        headers = [
            "Rank", 
            "User ID", 
            "First Name", 
            "Username", 
            "Total Invited", 
            "Active Invited", 
            "Join Date"
        ]
        
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            
        # Write data
        for row_num, member in enumerate(members, start=2):
            ws.cell(row=row_num, column=1, value=row_num - 1)
            ws.cell(row=row_num, column=2, value=str(member.user_id))
            ws.cell(row=row_num, column=3, value=member.user.first_name)
            ws.cell(row=row_num, column=4, value=f"@{member.user.username}" if member.user.username else "")
            ws.cell(row=row_num, column=5, value=member.total_invited)
            ws.cell(row=row_num, column=6, value=member.active_invited)
            ws.cell(row=row_num, column=7, value=member.joined_at.strftime("%Y-%m-%d %H:%M:%S") if member.joined_at else "")
            
        # Adjust column widths (basic estimation)
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
            
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
